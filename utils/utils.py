from openpyxl import load_workbook

def read_excel(filename:str) -> dict:
    """
    Read the Excel file and return a dictionary containing values for each row
    
    Args:
        filename: file path
    Return: 
        lis of dictionary
    """

    wb = load_workbook(filename)
    ws = wb.active

    max_rows = ws.max_row

    rows = []
    for row in range(2, max_rows + 1):

        current_row = {
            "first_name": ws[f"A{row}"].value,
            "last_name": ws[f"B{row}"].value,
            "company_name": ws[f"C{row}"].value,
            "role_in_company": ws[f"D{row}"].value,
            "address": ws[f"E{row}"].value,
            "email": ws[f"F{row}"].value,
            "phone_number": ws[f"G{row}"].value,
        }

        rows.append(current_row)
    
    return rows